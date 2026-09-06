from __future__ import annotations

import io
import re
import warnings
from datetime import datetime
from urllib.parse import urlencode

from openpyxl import load_workbook

from app.wind_agents.base import AgentFinding, BaseWindAgent


BASE_URL = "https://dati.terna.it"
DOWNLOAD_CENTER_URL = "https://dati.terna.it/download-center"
EXCEL_EXPORT_URL = "https://dati.terna.it/api/sitecore/dati/downloadcenter/records"
DATASET = "FER"
VIEW_BY = "Region"
DB = "enti"
DEFAULT_PAGE_SIZE = 203
TARGET_SOURCES = {"eolico", "wind"}


class TernaEconnextionWindAgent(BaseWindAgent):
    """Wind aggregate intelligence from Terna Econnextion.

    This source is intentionally never emitted as a project source: rows are
    regional connection-request aggregates and cannot be converted into named
    projects, developers or construction scopes.
    """

    agent_name = "institutional_watch"
    source_name = "Terna Econnextion - Eolico"
    base_url = DOWNLOAD_CENTER_URL

    @staticmethod
    def _clean(value: object) -> str | None:
        if value is None:
            return None
        text = " ".join(str(value).replace("\xa0", " ").split()).strip()
        return text or None

    @staticmethod
    def _to_float(value: object) -> float | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(" ", "")
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".") if text.rfind(",") > text.rfind(".") else text.replace(",", "")
        elif "," in text:
            text = text.replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None

    @staticmethod
    def _to_int(value: object) -> int | None:
        if value is None:
            return None
        try:
            return int(float(str(value).replace(",", ".")))
        except ValueError:
            return None

    @staticmethod
    def _key(value: str | None) -> str:
        return re.sub(r"[^a-z0-9]+", "-", (value or "").lower()).strip("-")

    def _headers(self, referer: str = DOWNLOAD_CENTER_URL) -> dict[str, str]:
        return {
            "User-Agent": "Wind-Radar-Agent/0.6",
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
            "Accept-Language": "it-IT,it;q=0.9,en;q=0.8",
            "Referer": referer,
            "Origin": BASE_URL,
        }

    def _bootstrap(self) -> None:
        response = self.session.get(
            DOWNLOAD_CENTER_URL,
            headers={
                "User-Agent": "Wind-Radar-Agent/0.6",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "it-IT,it;q=0.9,en;q=0.8",
            },
            timeout=90,
            allow_redirects=True,
        )
        response.raise_for_status()

    @staticmethod
    def _candidate_months(lookback: int = 8) -> list[tuple[int, int]]:
        now = datetime.now()
        year, month = now.year, now.month
        out: list[tuple[int, int]] = []
        for _ in range(lookback):
            out.append((year, month))
            month -= 1
            if month == 0:
                month = 12
                year -= 1
        return out

    @staticmethod
    def _build_url(year: int, month: int) -> str:
        params = {
            "f": "xlsx",
            "filterDataset": DATASET,
            "filterViewBy": VIEW_BY,
            "filterYear": str(year),
            "filterMonth": str(month),
            "orderByColumn": "Potenza (MW)",
            "orderByDir": "desc",
            "db": DB,
            "pageSize": str(DEFAULT_PAGE_SIZE),
        }
        return f"{EXCEL_EXPORT_URL}?{urlencode(params)}"

    def _download(self, year: int, month: int) -> bytes:
        response = self.session.get(
            self._build_url(year, month),
            headers=self._headers(),
            timeout=90,
            allow_redirects=True,
        )
        response.raise_for_status()
        if not response.content.startswith(b"PK"):
            raise RuntimeError("Terna Econnextion export is not a valid XLSX")
        return response.content

    @classmethod
    def _rows(cls, content: bytes) -> list[dict]:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Workbook contains no default style.*",
                category=UserWarning,
            )
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            iterator = sheet.iter_rows(values_only=True)
            try:
                headers_raw = next(iterator)
            except StopIteration:
                return []
            headers = [cls._clean(value) for value in headers_raw]
            rows: list[dict] = []
            for raw in iterator:
                row = {
                    header: raw[index] if index < len(raw) else None
                    for index, header in enumerate(headers)
                    if header
                }
                if any(value is not None and str(value).strip() for value in row.values()):
                    rows.append(row)
            return rows
        finally:
            workbook.close()

    def _latest(self) -> tuple[int, int, list[dict]] | None:
        for year, month in self._candidate_months():
            try:
                rows = self._rows(self._download(year, month))
                if rows:
                    return year, month, rows
            except Exception:
                continue
        return None

    def fetch(self) -> list[AgentFinding]:
        self._bootstrap()
        latest = self._latest()
        if latest is None:
            raise RuntimeError("no recent Terna Econnextion FER export available")
        year, month, rows = latest

        findings: list[AgentFinding] = []
        for row in rows:
            region = self._clean(row.get("Regione"))
            source = self._clean(row.get("Fonte"))
            status = self._clean(row.get("Stato Connessione"))
            plant_type = self._clean(row.get("Tipo Impianto"))
            if not region or not source or source.lower() not in TARGET_SOURCES:
                continue
            power_mw = self._to_float(row.get("Potenza (MW)"))
            if power_mw is None:
                continue
            practices = self._to_int(row.get("Numero Pratiche")) or 0
            external_id = (
                f"TERNA-ECONNEXTION-WIND-{year}-{month:02d}-"
                f"{self._key(region)}-{self._key(source)}-{self._key(status or 'nd')}"
            )
            findings.append(
                AgentFinding(
                    external_id=external_id,
                    source_name=self.source_name,
                    source_url=f"{DOWNLOAD_CENTER_URL}#{external_id}",
                    title=f"Terna Econnextion - {region} - {source} - {status or 'n.d.'}",
                    finding_type="market_aggregate",
                    payload={
                        "region": region,
                        "source": source,
                        "plant_type": plant_type,
                        "connection_status": status,
                        "power_mw": power_mw,
                        "number_of_requests": practices,
                        "reference_year": year,
                        "reference_month": month,
                        "excel_export_url": self._build_url(year, month),
                        "sector": "eolico",
                        "is_aggregated_market_intelligence": True,
                        "project_specific": False,
                        "execution_scope": None,
                        "source_grade_ceiling": "A1 aggregate",
                        "evidence_layer": "market_intelligence",
                        "aggregation_guard": "Regional connection-request aggregates must never be converted into named projects, developers or execution awards.",
                        "source_adapter_origin": "pv_agent_mvp/terna_econnextion.py",
                    },
                )
            )
        return findings
